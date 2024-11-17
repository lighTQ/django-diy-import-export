#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
@author: HighWay.Li
@contact:byamian@gmail.com
@version: 1.0.0
@license: Apache Licence
@file: excel_util.py
@time: 11/17/24 AM10:44
"""
# -*- coding: utf-8 -*-

import tempfile
from zipfile import ZipFile, ZIP_DEFLATED

from django.http import HttpResponse, StreamingHttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Side, Border, PatternFill, Protection
from openpyxl.writer.excel import ExcelWriter


class ExcelUtils:
    """excel处理类"""

    def __init__(self, file_name, header: list = None, data: list = None, sheet_name="Sheet"):
        """
        :param sheet_name: sheet_name
        :param header: 表头
        :param data: 数据
        """
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.header = header
        self.data = data

    def draw(self, header_styles=None, data_styles=None):
        """excel数据、样式设置并返回（导出）"""
        if data_styles is None:
            data_styles = {}
        if header_styles is None:
            header_styles = {}
        wb, sheet = self.excel_data_deal()
        self.excel_header_style(sheet, **header_styles)
        self.excel_data_style(sheet, **data_styles)
        # return self.excel_file_response(save_virtual_workbook(wb), self.file_name)
        return self.excel_stream_response(wb, self.file_name)

    def excel_data_deal(self):
        """创建excel并加入数据"""
        wb = Workbook()
        sheet = wb.create_sheet(self.sheet_name) if self.sheet_name != "Sheet" else wb.active
        if self.header is not None:
            sheet.append(self.header)

        for i, r_data in enumerate(self.data):
            for j, c_data in enumerate(r_data):
                sheet.cell(row=i + 2, column=j + 1, value=c_data)

        return wb, sheet

    def excel_header_style(self, sheet, **kwargs):
        """
        表头样式设置
        rows_width示例：设置E列宽度为30 {"E": 30}
        """
        sheet.row_dimensions[1].height = 20
        for i, r in enumerate(sheet[1]):
            width = len(r.value) * 3
            # 获取列字母
            column_letter = [chr(i + 65) if i <= 25 else 'A' + chr(i - 26 + 65)][0]
            # 列宽, 最大为50
            sheet.column_dimensions['{}'.format(column_letter)].width = width if width < 50 else 50
            self.write_excel_cell(r, **kwargs)

        if kwargs.get("rows_width"):
            for k, v in kwargs["rows_width"].items():
                sheet.column_dimensions[k].width = v

    def excel_data_style(self, sheet, **kwargs):
        """数据样式设置"""
        kwargs["fill"] = kwargs.get("fill", {})
        data_len = len(self.data)
        for i in range(data_len):
            for j, _ in enumerate(sheet.column_dimensions):
                kwargs["fill"]["fill_color"] = "00FFFF00" if j % 2 == 0 else "00FFFFFF"
                self.write_excel_cell(sheet.cell(row=i + 2, column=j + 1), **kwargs)

    @staticmethod
    def write_excel_cell(cell_obj, value=None, **kwargs):
        """
        单元格样式设置
        :param cell_obj: 单元格
        :param value: 单元格内容
        :alignment horizontal: 水平位置 "general", "left", "center", "right",
        :alignment vertical: 垂直位置 "top", "center", "bottom", "justify", "distributed"
        :alignment wrap: 是否换行
        :alignment shrink: 是否缩小填充
        :font size: 字体大小
        :font bold: 字体是否加粗，True/False
        :border side_style: 边框样式 默认thin
        :border side_color: 边框颜色 默认000000
        :fill fill_type: 填充样式 默认solid
        :fill fill_color: 填充颜色 默认无(白色)
        """
        if value is not None:
            cell_obj.value = value
        alignment = kwargs.get("alignment", {})
        font = kwargs.get("font", {})
        border = kwargs.get("border", {})
        fill = kwargs.get("fill", {})
        number_format = kwargs.get("number_format", "General")
        protection = kwargs.get("protection", {})

        # 位置：对齐方式等, 默认居中
        cell_obj.alignment = Alignment(horizontal=alignment.get("horizontal", "center"),
                                       vertical=alignment.get("vertical", "center"),
                                       wrapText=alignment.get("wrap", False),
                                       shrinkToFit=alignment.get("shrink", False))
        # 字体：字号、字体颜色、下划线等， 默认字体大小为10
        cell_obj.font = Font(size=font.get("size", None),
                             bold=font.get("bold", False))
        # 边框样式，默认黑色细线
        side = Side(style=border.get("side_style", "thin"),
                    color=border.get("side_color", "000000"))
        cell_obj.border = Border(left=side, right=side, top=side, bottom=side)
        # 填充：填充色、填充类型等，默认白色填充
        cell_obj.fill = PatternFill(fill_type=fill.get("fill_type", "solid"),
                                    fgColor=fill.get("fill_color", "00FFFFFF"))
        # 数据格式
        cell_obj.number_format = number_format
        # 写保护
        cell_obj.protection = Protection(locked=protection.get("locked", True),
                                         hidden=protection.get("hidden", False))

    @staticmethod
    def excel_file_response(content, file_name):
        """
        导出已处理的excel文件
        :param content: 文件内容
        :param file_name: 文件名
        """
        response = HttpResponse(content, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=%s' % file_name.encode().decode('latin-1')
        return response

    def excel_stream_response(self, wb, file_name):
        """返回流式数据"""
        response = StreamingHttpResponse(self.wb_iterator(wb), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response

    @staticmethod
    def wb_iterator(workbook, chunk_size=512):
        """重写save_virtual_workbook方法，适配文件流"""
        tmp = tempfile.TemporaryFile()
        archive = ZipFile(tmp, 'w', ZIP_DEFLATED, allowZip64=True)

        writer = ExcelWriter(workbook, archive)
        writer.save()

        tmp.seek(0)
        while True:
            virtual_workbook_content = tmp.read(chunk_size)
            if virtual_workbook_content:
                yield virtual_workbook_content
            else:
                tmp.close()
                break