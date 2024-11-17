#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
@author: HighWay.Li
@contact:byamian@gmail.com
@version: 1.0.0
@license: Apache Licence
@file: excel_opt.py
@time: 11/17/24 PM5:21
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from django.db.models.query import QuerySet


class ExcelExporter:
    def __init__(self, headers=None, field_types=None, text_columns=None):
        """
        初始化ExcelExporter对象
        :param headers: 表头名称列表
        :param field_types: 字段类型字典, 例如: {'field1': str, 'field2': int}
        :param text_columns: 需要设置为文本格式的列名列表
        """
        self.headers = headers
        self.field_types = field_types
        self.text_columns = text_columns if text_columns is not None else []

    def export(self, data, file_path):
        """
        导出数据到Excel文件
        :param data: pandas DataFrame 或 Django QuerySet 对象
        :param file_path: 导出的Excel文件路径
        """
        if isinstance(data, QuerySet):
            data = pd.DataFrame(list(data.values()))

        if self.headers:
            data.columns = self.headers

        if self.field_types:
            for field, dtype in self.field_types.items():
                data[field] = data[field].astype(dtype)

        wb = Workbook()
        ws = wb.active

        # 定义文本样式
        text_style = NamedStyle(name="text_style")
        text_style.number_format = '@'

        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # 如果是需要设置为文本格式的列，应用文本样式
                if ws.cell(row=1, column=c_idx).value in self.text_columns:
                    cell.style = text_style

        wb.save(file_path)


if __name__ == '__main__':

    # 使用示例

    df = pd.DataFrame({
        'name': ['Alice', 'Bob', 'Charlie'],
        'id': ['00123', '00456', '00789']
    })
    # 或者一个QuerySet对象（通过Django ORM查询得到）
    # queryset = Data.objects.all()
    # 定义表头、字段类型和需要设置为文本格式的列
    headers = ['Name', 'ID']
    field_types = {'Name': str, 'ID': str}
    text_columns = ['ID']

    exporter = ExcelExporter(headers=headers, field_types=field_types, text_columns=text_columns)
    exporter.export(df, 'output.xlsx')

    # 对于 QuerySet 对象，使用类似的方式：
    # exporter.export(queryset, 'output.xlsx')
