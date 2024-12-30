#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
@author: HighWay.Li
@contact:byamian@gmail.com
@version: 1.0.0
@license: Apache Licence
@file: example_view.py
@time: 12/29/24 PM2:02
"""
import io

import openpyxl
import pandas as pd
from django_filters.rest_framework.backends import DjangoFilterBackend
from openpyxl.styles.named_styles import NamedStyle
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse

from api.models.models import ExampleModel, ExampleModelSerializer
from api.views.basic_view import BasicView


def underscore_to_camelcase(value):
    parts = value.split('_')
    return parts[0] + ''.join(word.capitalize() for word in parts[1:])


class BatchUpdateView(BasicView):
    queryset = ExampleModel.objects.all()
    serializer_class = ExampleModelSerializer

    filter_backends = (DjangoFilterBackend,)
    filterset_fields = "__all__"
    # ordering_fields = "__all__"
    ordering_fields = ['id', 'transport_no']

    @action(detail=False, methods=['post'])
    def batchUpdate(self, request, *args, **kwargs):
        data = request.data
        transport_no = data['transport_no']
        lines = data['lines']

        # 批量更新
        for item in lines:
            ExampleModel.objects.filter(id=item['id']).update(**item)
        # 根据 transport_no 查询数据库

        # 将列名转换为列表
        columns=['doc_no',]
        column_list = columns.split(',')
        # 查询数据库并只选择指定列
        queryset = ExampleModel.objects.filter(transport_no=transport_no).only(*column_list)
        # queryset = ExampleModel.objects.filter(transport_no=transport_no)
        if not queryset.exists():
            return Response({"msg": "没有找到对应的记录"}, status=status.HTTP_404_NOT_FOUND)

        # 反序列化生成 DataFrame
        serializer = ExampleModelSerializer(queryset, many=True)
        df = pd.DataFrame(serializer.data)

        # 判断 DataFrame 是否包含空值
        if df.isnull().values.any():
            return Response({"msg": "数据包含空值，不能生成 Excel 文件"}, status=status.HTTP_400_BAD_REQUEST)

        # 生成 Excel 文件
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="export.xlsx"'

        # 转换表头为首字母小写的驼峰命名
        df.columns = [underscore_to_camelcase(col) for col in df.columns]

        # # 写入 Excel 文件1
        # df.to_excel(response, index=False)
        # return response


        # 创建一个字节流来保存 Excel 文件2
        output = io.BytesIO()
        # 使用 pandas 和 xlsxwriter 写入 Excel 文件
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title='Data'
        text_style = NamedStyle(name="text_style", number_format="@")
        workbook.add_named_style(text_style)

        # 写入表头
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num, value=column_title)
            cell.style = text_style

        # 写入数据
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)
                cell.style = text_style
                # 保存到字节流
        workbook.save(output)
        output.seek(0)

        # 创建响应
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
        return response



