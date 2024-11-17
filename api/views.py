#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
@author: HighWay.Li
@contact:byamian@gmail.com
@version: 1.0.0
@license: Apache Licence
@file: views.py
@time: 11/15/24 PM7:02
"""
from http.client import responses
from xml.dom import ValidationErr

import pandas as pd
from django.http import HttpResponse, JsonResponse
from numpy.distutils.misc_util import blue_text
from numpy.f2py.crackfortran import publicpattern
from rest_framework import status, viewsets
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.templatetags.rest_framework import add_query_param
from rest_framework.viewsets import GenericViewSet, ModelViewSet
from rest_framework.request import Request

from api.forms import ImportFileForm
from api.models import PersonSerializer, Person
from api.serializer import GoodsSerializer, PublishSerializer, PublishModelSerializer
from goods.models import Goods, Publish, Books
from rest_framework.views import View
import xlwt


@api_view(['GET'])
def get_data0(request):
    goods = {"name": "测试商品", "price": 100.0}
    return Response(data=goods)


@api_view(['GET', 'POST'])
def goods_list(request):
    if request.method == 'GET':
        goods = Goods.objects.all()
        serializer = GoodsSerializer(goods, many=True)
        return Response(data=serializer.data)
    if request.method == 'POST':
        serializer = GoodsSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(data=serializer.data, status=status.HTTP_201_CREATED)


@api_view(['GET', 'PUT', 'DELETE'])
def goods_detail(request, id):
    try:
        goods = Goods.objects.get(id=id)
    except Goods.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = GoodsSerializer(goods)
        return Response(data=serializer.data)
    if request.method == 'PUT':
        serializer = GoodsSerializer(goods, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        return Response(data=serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    if request.method == 'DELETE':
        goods.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class PublishViewSet(ModelViewSet):
    queryset = Publish.objects.all()
    serializer_class = PublishSerializer


## CBV 源码执行流程
class Book(View):

    def dispatch(self, request, *args, **kwargs):
        # add code

        responses = super().dispatch(request, *args, **kwargs)
        return responses

    def get(self, request, *args, **kwargs):
        return HttpResponse('get response')

    def post(self, request, *args, **kwargs):
        return HttpResponse('post response')


from rest_framework.views import APIView
from rest_framework.response import Response

# class BookAPIView(APIView):
#     def get(self, request, *args, **kwargs):
#         return Response("APIView Response")


from api.serializer import BooksSerializer


class BookAPIView(APIView):
    def get(self, request, *args, **kwargs):
        book_list = Books.objects.all()
        # 序列化过程，把QuerySet转换成字典
        # 第一个参数就是要序列化的对象，如果序列化多条数据，则many=True一定要加上
        serializer = BooksSerializer(book_list, many=True)
        print(serializer.data)
        return Response(serializer.data)


# 查看单个的接口
class BookDetailAPIView(APIView):
    # def get(self, request,name,price):
    #     book_list = Books.objects.filter(price=price,name=name)
    #     ser = BooksSerializer(book_list, many=True)
    #     print(ser.data)
    #     return Response(ser.data)

    def __init__(self, **kwargs):
        super().__init__(kwargs)
        self.data = None

    def get(self, request, name):
        book = Books.objects.get(name=name)
        serializer = BooksSerializer(book)
        print(serializer.data)
        return Response(serializer.data)

    def delete(self, request, name):
        res = Books.objects.filter(name=name).delete()
        if res:
            return Response(status=status.HTTP_204_NO_CONTENT)

    def put(self, request, name):
        book = Books.objects.filter(name=name).first()
        # 反序列化
        ser = BooksSerializer(instance=book, data=request.data)

        if ser.is_valid():
            print(ser.validated_data)
            ser.save()
            return Response(ser.data)

    def post(self, request, name):
        ser = BooksSerializer(data=request.data)
        if ser.is_valid():
            ser.save()
            return Response(ser.data)
        return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)

    # def export_users_xls(request):  #     print(request.data)  #     response = HttpResponse(content_type='application/ms-excel')  #     response['Content-Disposition'] = 'attachment; filename="books.xlsx"'  #  #     wb = xlwt.Workbook(encoding='utf-8')  #     ws = wb.add_sheet('Users', cell_overwrite_ok=True)  #  #     row_num = 0  #     font_style = xlwt.XFStyle()  #     font_style.font.bold = True  #     columns = ['id', 'name', 'price', 'publish']  #     for col_num in range(len(columns)):  #         ws.write(row_num, col_num, columns[col_num], font_style)  #     font_style = xlwt.XFStyle()  #     rows = Books.objects.all().values_list('id', 'name', 'price', 'publish')  #     for row in rows:  #         row_num += 1  #         for col_num in range(len(row)):  #             ws.write(row_num, col_num, row[col_num], font_style)  #     wb.save(response)  #     return response


class PublishView(APIView):
    queryset = Publish.objects.all()
    serializer_class = PublishSerializer
    def get(self, request, *args, **kwargs):
        # ser = self.serializer_class(self.self.queryset, many=True)
        ser = PublishModelSerializer(self.queryset, many=True)
        print(ser.data)
        return Response(ser.data)

    def post(self, request, *args, **kwargs):
        ser = self.serializer_class(data=request.data)
        if ser.is_valid():
            ser.save()
            return Response(ser.data)
        else:
            print(ser.errors.get('name'))
            return Response({'msg': '出错了！！'})


class ExportExcelView(APIView):
    def post(self, request, *args, **kwargs):
        ser = PublishSerializer(request.data)
        name = ser.data.get('name')
        addr = ser.data.get('address')
        print(ser.data)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="publish_single.xlsx"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Publish', cell_overwrite_ok=True)

        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        columns = ['id', 'name', 'addr']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        font_style = xlwt.XFStyle()
        rows = Publish.objects.filter(name=name, addr=addr).values_list()

        sql = "select id ,name,addr from goods_publish where name = '{0}' and addr ='{1}'".format(name, addr)
        print('sql:{0}'.format(sql))
        # rows = Publish.objects.raw(sql)
        # rows = Books.objects.all().values_list('id', 'name', 'price', 'publish')

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
        wb.save(response)

        return response


from django.shortcuts import render
from openpyxl import load_workbook


class ImportExcelView(APIView):
    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            excel_file = request.FILES['file']
            wb = load_workbook(excel_file)
            ws = wb.active
            bulk_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                name, addr = row
                bulk_list.append(Publish(name=name, addr=addr))
            exist_name_list = set([x.name for x in Publish.objects.all()])
            new_data_list = []
            for y in bulk_list:
                if y.name not in exist_name_list:
                    new_data_list.append(y)

            Publish.objects.bulk_create(new_data_list)
        return Response(status=status.HTTP_201_CREATED)


"""
POST:
localhost:8000/import_file/
param:
Key Value
file File A person.xlsx
Sex Text female
"""


class ImportFileView(APIView):

    def get(self, request, *args, **kwargs):
        person_list = Person.objects.all()
        ser = PersonSerializer(person_list, many=True)

        return Response(ser.data)

    parser_class = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file = request.FILES['file']
        add_params = request.data.get('sex')
        if file and add_params:
            try:
                # read excel file
                df = pd.read_excel(file)
                # check and validate data
                valid_data = df[df['sex'] == add_params]
                if not valid_data.empty:
                    # 讲校验之后的数据保存到数据库
                    for _, row in valid_data.iterrows():
                        ser = PersonSerializer(data=dict(row))
                        if ser.is_valid():
                            ser.save()
                        else:
                            return Response({'error': 'data validate failure'}, status=status.HTTP_400_BAD_REQUEST)
                    return Response({'success': '文件上传并校验成功'}, status=status.HTTP_201_CREATED)
                else:
                    return Response({'error': 'data validate failure'}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        else:
            return Response({'error': '文件或参数缺失'}, status=status.HTTP_400_BAD_REQUEST)


"""
POST:
localhost:8000/export_file/
param:
{
    "query_param": {
        "name": "",
        "sex": "female"
    }
}
"""


class ExportFileView(APIView):
    parser_class = [JSONParser]

    def post(self, request, *args, **kwargs):
        try:
            query_param = request.data.get('query_param', {})
            qs = Person.objects.all()
            if 'name' in query_param and query_param['name'] != '':
                qs = qs.filter(name__icontains=query_param['name'])
            if 'sex' in query_param:
                qs = qs.filter(sex__icontains=query_param['sex'])
            data = list(qs.values())
            ser = PersonSerializer(data=data, many=True)
            # 序列化数据
            # ser = PersonSerializer(data=list(qs), many=True)
            if ser.is_valid():
                # data transform
                df = pd.DataFrame(ser.data)
                print(df)
                df['salary'] = df['age'] * 100
                # 将数据写入Excel文件
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=exported_data.xlsx'
                with pd.ExcelWriter(response, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                return response
            else:
                return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


